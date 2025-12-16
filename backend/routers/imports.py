from fastapi import APIRouter, UploadFile, File, Form, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.ext.asyncio import AsyncSession
from backend.database import get_db
from backend import models, crud, schemas, auth
from datetime import datetime
import io
import csv
import openpyxl
import re

router = APIRouter(prefix="/import", tags=["import"])

HEADER = ["DESCRICAO", "CATEGORIA", "FORNECEDOR_NOME", "FORNECEDOR_CNPJ", "DATA_COMPRA", "VALOR", "NUMERO_NOTA", "NUMERO_SERIE", "ATIVO_FIXO", "OBSERVACOES"]
ROW = ["NOTEBOOK DELL", "INFORMATICA", "DELL COMPUTADORES", "12.345.678/0001-99", "10/12/2024", "3450.99", "NF123456", "ABC123", "ATV-12345", "ITEM IMPORTADO VIA EXEMPLO"]

@router.get("/example-csv")
async def get_example_csv():
    output = io.StringIO()
    writer = csv.writer(output, delimiter=';', quoting=csv.QUOTE_MINIMAL)
    writer.writerow(HEADER)
    writer.writerow(ROW)
    output.seek(0)

    mem = io.BytesIO()
    mem.write(output.getvalue().encode('utf-8-sig'))
    mem.seek(0)

    return StreamingResponse(
        mem,
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=exemplo_importacao.csv"}
    )

@router.get("/example-xlsx")
async def get_example_xlsx():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Importacao"

    ws.append(HEADER)
    ws.append(ROW)

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)

    return StreamingResponse(
        mem,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=exemplo_importacao.xlsx"}
    )

@router.post("/upload")
async def upload_import(
    branch_id: int = Form(...),
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    # Permission Check
    if current_user.role == models.UserRole.AUDITOR:
        raise HTTPException(status_code=403, detail="Auditores não podem importar itens")

    # Branch Permission
    if current_user.role == models.UserRole.OPERATOR and not current_user.all_branches:
        allowed = [b.id for b in current_user.branches]
        if current_user.branch_id: allowed.append(current_user.branch_id)
        if branch_id not in allowed:
             raise HTTPException(status_code=403, detail="Sem permissão para esta filial")

    # Read File
    contents = await file.read()
    filename = file.filename.lower()

    rows = []

    try:
        if filename.endswith('.csv'):
            # Decode with BOM handling
            content_str = contents.decode('utf-8-sig')
            # Use semi-colon delimiter
            reader = csv.DictReader(io.StringIO(content_str), delimiter=';')
            # Normalize headers (upper, strip)
            reader.fieldnames = [f.strip().upper() for f in reader.fieldnames] if reader.fieldnames else []
            rows = list(reader)

        elif filename.endswith(('.xls', '.xlsx')):
            wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
            ws = wb.active
            data_rows = list(ws.iter_rows(values_only=True))
            if data_rows:
                headers = [str(h).strip().upper() if h else f"COL_{i}" for i, h in enumerate(data_rows[0])]
                for row_values in data_rows[1:]:
                    # Create dict only if row has data
                    if any(row_values):
                        # Handle length mismatch if any
                        row_dict = dict(zip(headers, row_values))
                        rows.append(row_dict)
        else:
            raise HTTPException(status_code=400, detail="Formato não suportado. Use .csv ou .xlsx")

    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao ler arquivo: {str(e)}")

    success_count = 0
    errors = []

    for index, row in enumerate(rows):
        try:
            # Map Row Keys (handle potential variations if needed, but assuming strict header)

            # 1. Date
            raw_date = row.get("DATA_COMPRA")
            p_date = None
            if raw_date:
                if isinstance(raw_date, datetime):
                    p_date = raw_date
                else:
                    date_str = str(raw_date).strip()
                    for fmt in ["%d/%m/%Y", "%Y-%m-%d", "%d-%m-%y"]:
                        try:
                            p_date = datetime.strptime(date_str, fmt)
                            break
                        except:
                            pass
            if not p_date:
                p_date = datetime.now() # Fallback

            # 2. Value
            raw_val = row.get("VALOR", 0)
            invoice_value = 0.0
            if isinstance(raw_val, (int, float)):
                invoice_value = float(raw_val)
            else:
                val_str = str(raw_val).strip().replace("R$", "").strip()
                if ',' in val_str and '.' in val_str:
                     val_str = val_str.replace('.', '').replace(',', '.')
                elif ',' in val_str:
                     val_str = val_str.replace(',', '.')
                try:
                    invoice_value = float(val_str)
                except:
                    invoice_value = 0.0

            # 3. Supplier
            supplier_name = row.get("FORNECEDOR_NOME")
            supplier_cnpj = row.get("FORNECEDOR_CNPJ")
            supplier_id = None

            if supplier_cnpj and str(supplier_cnpj).strip():
                clean_cnpj = re.sub(r'\D', '', str(supplier_cnpj))
                if clean_cnpj:
                    sup = await crud.get_supplier_by_cnpj(db, clean_cnpj)
                    if not sup and supplier_name:
                        # Create Supplier
                        sup_in = schemas.SupplierCreate(name=str(supplier_name), cnpj=clean_cnpj)
                        sup = await crud.create_supplier(db, sup_in)
                    if sup:
                        supplier_id = sup.id

            # 4. Category
            cat_name = str(row.get("CATEGORIA", "")).strip().upper()
            cat_id = None
            if cat_name:
                c = await crud.get_category_by_name(db, cat_name)
                if c:
                    cat_id = c.id
                else:
                    # Create category if missing?
                    # Prompt didn't specify, but it's better UX.
                    # Given the "Example file" structure, user expects it to work.
                    cat_in = schemas.CategoryCreate(name=cat_name)
                    c = await crud.create_category(db, cat_in)
                    cat_id = c.id

            # 5. Description & Others
            desc = str(row.get("DESCRICAO", "")).upper()
            if not desc or desc == "NONE": # Handle xlsx None
                continue # Skip empty rows

            item_in = schemas.ItemCreate(
                description=desc,
                category=cat_name,
                category_id=cat_id,
                purchase_date=p_date,
                invoice_value=invoice_value,
                invoice_number=str(row.get("NUMERO_NOTA", "")),
                branch_id=branch_id,
                supplier_id=supplier_id,
                serial_number=str(row.get("NUMERO_SERIE", "")),
                fixed_asset_number=str(row.get("ATIVO_FIXO", "")),
                observations=str(row.get("OBSERVACOES", "")),
                responsible_id=current_user.id,
                status=models.ItemStatus.APPROVED if current_user.can_import else models.ItemStatus.PENDING
            )

            log_msg = "Item importado via arquivo. Auto-aprovado por permissão." if current_user.can_import else "Item importado via arquivo. Aguardando aprovação."
            await crud.create_item(db, item_in, action_log=log_msg)
            success_count += 1

        except Exception as e:
            errors.append(f"Linha {index+2}: {str(e)}")

    return {"success": success_count, "errors": errors}
